from PySimpleGUI.PySimpleGUI import Cancel, Ok, PopupOKCancel, popup
from openpyxl import Workbook
import PySimpleGUI as sg

def CriaLinhaTitulo(page):
    page.append(['ID', 'Marca', 'Produto', 'Perfil', 'Sistema', 'NomeDis', 'DataCriacao', 'Tipo', 'DataGarantia', 'Destino', 'Pedido', 'Status', 'Ícone'])

def CriaLinhas(page, quantidade_de_linhas, id, produto, perfil, sistema, nomedis, datacriacao, tipo, datagarantia, destino, pedido, status, icone):
     contador = 0
     converter = int(quantidade_de_linhas)
     while (contador < converter ):
         id = int(id) + 1
         page.append([id,'', produto, perfil, sistema,nomedis, datacriacao, tipo,datagarantia, destino, pedido, status,icone])
         contador += 1
     
# Criada a planilha
book = Workbook()
book.create_sheet('Import', 0)

# Selecionando a planilha
page_import = book['Import']

print(book.sheetnames)

# Janelas e estilos(layouts)

def janela_login():
    sg.theme('DefaultNoMoreNagging')
    layout = [
        [sg.Text('Usuario')],
        [sg.Input(key = 'usuario')],
        [sg.Text('Senha')],
        [sg.Input(key = 'senha',password_char='*')],
        [sg.Button('Entrar'), sg.Button('Sair')]

    ]

    return sg.Window('Login', layout=layout, finalize=True)

def janela_dados_da_planilha():
    sg.theme('DefaultNoMoreNagging')
    layout = [

        [sg.Text('Quantidade de Linha')],
        [sg.Input(int(),key='quantidade_de_linhas')],

        [sg.Text('ID')],
        [sg.Input(int('3870'),key='id')],

        [sg.Text('Produto')],
        [sg.Combo(['Cadeado CLIQ G55','Chave Temporary Key', 'Cilindro CLIQ'],size=(43,1), key='produto')],

        [sg.Text('Pefil')],
        [sg.Input(key='perfil')],

        [sg.Text('Sistema')],
        [sg.Combo(['Vivo Teste','Claro Teste','Nextel Teste','Enel Teste','GTS Teste'],size=(43,1),key='sistema')],
       
        [sg.Text('Nomedis')],
        [sg.Input(key='nomedis')],

        [sg.Text('DataCriacao')],
        [sg.Input(key='datacriacao')],

        [sg.Text('Tipo')],
        [sg.Input(key='tipo')],

        [sg.Text('DataGarantia')],
        [sg.Input(key='datagarantia')],

        [sg.Text('Destino')],
        [sg.Input(key='destino')],

        [sg.Text('NumeroPedido')],
        [sg.Input(key='numeropedido')],

        [sg.Text('Status')],
        [sg.Input(key='status')],

        [sg.Text('Icone')],
        [sg.Input(key='icone')],

        [sg.Button('Voltar'), sg.Button('Salvar'),sg.Button('Importar')]
    ]


    return sg.Window("Dados", layout=layout, finalize=True)


# Criar as janelas iniciais
janela1, janela2 = janela_login(), None

usuario_jean, senha_jean = 'Jean', 'Cerveja'
usuario_marco, senha_marco = 'Boss', 'Jovemainda'
   
# Criando nosso loop de leitura de dados
while True:
    window,event,values = sg.read_all_windows()
    if window == janela1 and event == sg.WIN_CLOSED:
        break
    if window == janela1 and event == 'Sair':
        break
    if window == janela1 and event == 'Entrar':
        if values['usuario'] == usuario_jean and values['senha'] == senha_jean or values['usuario'] == usuario_marco and values['senha'] == senha_marco:
            popup('Welcome, My World')
            janela2 = janela_dados_da_planilha()
            janela1.hide()
        else:
            popup('Senha ou Usuario incorretos')
            janela2 = ''
    #Comecando a logica Janela 2
    if window == janela2 and event == sg.WIN_CLOSED:
        break
    if window == janela2 and event == 'Voltar':
        janela2.hide()
        janela1.un_hide()
    if window == janela2 and event == 'Salvar':
        sg.popup_ok_cancel('Verificar se tudo esta correto, antes de Importar',auto_close=False)  
        # CriaLinhaTitulo(page_import)
        # CriaLinhas(page_import, values['quantidade_de_linhas'],values['id'],values['produto'],values['perfil'],values['sistema'],values['nomedis'],values['datacriacao'],values['tipo'],values['datagarantia'],values['destino'],values['numeropedido'],values['status'],values['icone'])
    if window == janela2 and event == 'Importar':
        CriaLinhaTitulo(page_import)
        CriaLinhas(page_import, values['quantidade_de_linhas'],values['id'],values['produto'],values['perfil'],values['sistema'],values['nomedis'],values['datacriacao'],values['tipo'],values['datagarantia'],values['destino'],values['numeropedido'],values['status'],values['icone'])
        book.save('Importar.xlsx')
        
            